from quantum import quantum_random_numbers
from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
from classical import LCG

import math
from collections import Counter
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
import random
# this is for passsword
from classical import generate_classical_passwords
from quantum import generate_quantum_passwords
# end pswd

app = Flask(__name__)
CORS(app)

lcg = LCG(seed=98765)


@app.route("/")
def home():
    return "Backend running"

# classical api pswd
@app.route("/classical-passwords")
def classical_passwords():

    passwords = generate_classical_passwords(500)

    return jsonify({
        "type": "Classical Password Generator",
        "passwords": passwords
    })

# quantum api pswd
@app.route("/quantum-passwords")
def quantum_passwords():

    passwords = generate_quantum_passwords(500)

    return jsonify({
        "type": "Quantum Password Generator (IBM Hardware)",
        "passwords": passwords
    })



# ================= CLASSICAL RNG =================

@app.route("/classical/<int:count>")
def classical_rng(count):

    numbers = [lcg.generate() for _ in range(count)]

    return jsonify({
        "method": "Mersenne Twister",
        "random_numbers": numbers
    })
# ================= CLASSICAL LOTTERY =================

@app.route("/classical-lottery")
def classical_lottery():

    series_list = [
        "BA","BB","BC","BD","BE","BF","BG","BH",
        "BJ","BK","BL","BM","BN","BO","BP","BR"
    ]

    tickets = []

    for _ in range(500):

        series = random.choice(series_list)
        number = random.randint(100000, 999999)

        tickets.append(f"{series} {number}")

    return jsonify({
        "type": "Classical Kerala Lottery",
        "tickets": tickets
    })

# ================= QUANTUM RNG =================

@app.route("/quantum/<int:count>")
def quantum_rng(count):

    numbers = quantum_random_numbers(count)

    return jsonify({
        "method": "Quantum Random Number Generator (IBM Hardware)",
        "random_numbers": numbers
    })

# ================= QUANTUM LOTTERY =================
@app.route("/quantum-lottery")
def quantum_lottery():
    try:
        # ↓↓↓ Change from 1500 to something small for testing
        nums = quantum_random_numbers(90)   # 90 shots → ~30 tickets, much faster

        series_list = [
            "BA","BB","BC","BD","BE","BF","BG","BH",
            "BJ","BK","BL","BM","BN","BO","BP","BR"
        ]

        tickets = []

        for i in range(0, len(nums), 3):
            if i+2 >= len(nums):
                break
            series_idx = nums[i] % len(series_list)
            series = series_list[series_idx]

            # Use two 8-bit numbers → 0–65535 range, then map to 100000–999999
            number = (nums[i+1] * 256 + nums[i+2]) % 900000 + 100000

            tickets.append(f"{series} {number}")

        return jsonify({
            "type": "Quantum Kerala Lottery",
            "tickets": tickets
        })

    except Exception as e:
        import traceback
        print("Quantum lottery error:", str(e))
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500
# ================= EXCEL EXPORT =================

@app.route("/export-excel", methods=["POST"])
def export_excel():

    data = request.get_json()

    classical_numbers = data["classical"]
    quantum_numbers = data["quantum"]

    COUNT = len(classical_numbers)
    MAX_VALUE = 255

    # ---------- METRICS ----------

    def entropy(data):
        freq = Counter(data)
        total = len(data)
        H = 0

        for count in freq.values():
            p = count / total
            H -= p * math.log2(p)

        return round(H, 5)

    def chi_square(data):
        freq = Counter(data)
        expected = len(data) / (MAX_VALUE + 1)

        chi = 0

        for value in range(MAX_VALUE + 1):

            observed = freq.get(value, 0)

            chi += ((observed - expected) ** 2) / expected

        return round(chi, 5)

    # ---------- WORKBOOK ----------

    wb = Workbook()

    # ---------- RAW DATA ----------

    ws1 = wb.active
    ws1.title = "Random_Numbers"

    ws1["A1"] = "Classical"
    ws1["B1"] = "Quantum"

    for i in range(COUNT):

        ws1.cell(row=i+2, column=1, value=classical_numbers[i])
        ws1.cell(row=i+2, column=2, value=quantum_numbers[i])

    # ---------- ANALYSIS ----------

    ws2 = wb.create_sheet("Analysis")

    ws2.append(["Metric","Classical","Quantum"])

    ws2.append(["Total Numbers", COUNT, COUNT])

    ws2.append(["Entropy",
                entropy(classical_numbers),
                entropy(quantum_numbers)])

    ws2.append(["Chi-Square",
                chi_square(classical_numbers),
                chi_square(quantum_numbers)])

    # ---------- FREQUENCY ----------

    ws3 = wb.create_sheet("Frequency")

    ws3.append(["Value","Classical","Quantum"])

    classical_freq = Counter(classical_numbers)
    quantum_freq = Counter(quantum_numbers)

    for i in range(MAX_VALUE + 1):

        ws3.append([
            i,
            classical_freq.get(i,0),
            quantum_freq.get(i,0)
        ])

    # ---------- CHART ----------

    chart = BarChart()

    chart.title = "Distribution Comparison"

    data = Reference(ws3,
                     min_col=2,
                     min_row=1,
                     max_row=MAX_VALUE+2,
                     max_col=3)

    cats = Reference(ws3,
                     min_col=1,
                     min_row=2,
                     max_row=MAX_VALUE+2)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws3.add_chart(chart,"E2")

    # ---------- SAVE ----------

    file_path = "RNG_Report.xlsx"

    wb.save(file_path)

    return send_file(file_path, as_attachment=True)


# ================= RUN SERVER =================

if __name__ == "__main__":
    app.run(debug=True)